from urllib import request as rq
from bs4 import BeautifulSoup as bs
import openpyxl as xl


weatherUrl = 'http://www.kma.go.kr/weather/forecast/mid-term-rss3.jsp?stnId=133'

# rss 에서 가져왔으므로 xml 데이터
xmlData = rq.urlopen(weatherUrl)
# print(xmlData)

# 분석
parseData = bs(xmlData, 'html.parser')
# print(f'{parseData}')

# titles = parseData.find_all('title')
#
# for idx, title in enumerate(titles):
#     print(f'[{idx}] title : {title.string}')

modes = parseData.find_all('mode')
tmEfs = parseData.find_all('tmef')
wfs = parseData.find_all('wf')
tmns = parseData.find_all('tmn')
tmxs = parseData.find_all('tmx')
rnSts = parseData.find_all('rnst')

# 개수
# print(f'modes cnt : {len(modes)}')
# print(f'tmEfs cnt : {len(tmEfs)}')
# print(f'wfs cnt : {len(wfs)}')
# print(f'tmns cnt : {len(tmns)}')
# print(f'tmxs cnt : {len(tmxs)}')
# print(f'rmSts cnt : {len(rnSts)}')


for idx, mode in enumerate(modes):
     print(f'[{idx}] [{mode.string}] [{tmEfs[idx].string}] [{wfs[idx].string}] '
           f'[{tmns[idx].string}] [{tmxs[idx].string}] [{rnSts[idx].string}]')

modes = parseData.find_all('mode')
tmEfs = parseData.find_all('tmef')
wfs = parseData.find_all('wf')
tmns = parseData.find_all('tmn')
tmxs = parseData.find_all('tmx')
rnSts = parseData.find_all('rnst')


# ===== 엑셀 쓰기 ========
wb = xl.Workbook()
sheet = wb.active
sheet.title = '테스트'

col_names = ['mode', 'tmef', 'wf', 'tmn', 'tmx', 'rnst']
for seq, name in enumerate(col_names):
    sheet.cell(row=1, column=seq+1, value=name)

# faq_data = [('비밀번호 변경', '비밀번호 변경은 어떻게 해야 합니다.'), ('테스트 하는 방법', '테스트는 잘 해야 합니다.')]

for idx, mode in enumerate(modes):
    weatherData = [(mode.string, tmEfs[idx].string, wfs[idx].string, tmns[idx].string, tmxs[idx].string, rnSts[idx].string)]

    row_no = 2
    for n, rows in enumerate(weatherData):
        for seq, value in enumerate(rows):
            sheet.cell(row=row_no+n, column=seq+1, value=value)
            sheet.append([mode.string, tmEfs[idx].string, wfs[idx].string, tmns[idx].string, tmxs[idx].string,
                          rnSts[idx].string])
            wb.save("C:/chh_scraping/download/text2.xlsx")
            wb.close()




